"""
第二章 机理研究 - 数据分析与绘图（接入学术AI系统）
使用 academic_plot_style / statistical_analysis / scientific_visualization_agent
"""
import os, glob, sys
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib

# ══════════════════════════════════════
# 0. 接入学术AI系统
# ══════════════════════════════════════
sys.path.insert(0, os.path.dirname(__file__))

from academic_plot_style import (
    set_plot_style, save_figure, get_label, format_chemical,
    significance_stars, add_significance_bars,
    CHINESE_FONT, ENGLISH_FONT, CN_FONT_PROP, CN_FONT_PROP_BOLD,
    TABLEAU_10, OKABE_ITO, PHASE_COLORS,
)
from statistical_analysis import StatisticalAnalyzer
from scientific_visualization_agent import VisualizationAgent, StylePresets

# 应用学术样式
set_plot_style()
# 中文论文样式
StylePresets.apply('chinese')

# ── 输出目录 ──
OUT_DIR = os.path.expanduser('~/Desktop/第二章分析结果')
os.makedirs(OUT_DIR, exist_ok=True)

# ══════════════════════════════════════
# 1. 读取原始数据
# ══════════════════════════════════════
desktop = os.path.expanduser('~/Desktop')
xlsx_files = glob.glob(os.path.join(desktop, '*.xlsx'))
filepath = [f for f in xlsx_files if '机理' in f][0]

import openpyxl
wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb[wb.sheetnames[0]]

def cell(r, c):
    v = ws.cell(row=r, column=c).value
    return float(v) if v is not None else np.nan

# ── 分组定义 ──
M_GROUPS = ['MI', 'MII', 'MIII']
E_GROUPS = ['EI', 'EII', 'EIII']
ME_GROUPS = ['MEI', 'MEII', 'MEIII']
C_GROUPS = ['CI', 'CII']
ALL_GROUPS = M_GROUPS + E_GROUPS + ME_GROUPS + C_GROUPS

# ── 组标签 ──
LABELS = {
    'MI': 'MI (糖浆+砂+Cr+AN)', 'MII': 'MII (糖浆+砂+Cr)', 'MIII': 'MIII (糖浆+砂+AN)',
    'EI': 'EI (乳化油+砂+Cr+AN)', 'EII': 'EII (乳化油+砂+Cr)', 'EIII': 'EIII (乳化油+砂+AN)',
    'MEI': 'MEI (混合+砂+Cr+AN)', 'MEII': 'MEII (混合+砂+Cr)', 'MEIII': 'MEIII (混合+砂+AN)',
    'CI': 'CI (砂+Cr+AN)', 'CII': 'CII (Cr+AN)',
}

# 碳源分组标签（用于区分大组）
CARBON_SOURCE = {
    'MI': '糖浆', 'MII': '糖浆', 'MIII': '糖浆',
    'EI': '乳化油', 'EII': '乳化油', 'EIII': '乳化油',
    'MEI': '混合碳源', 'MEII': '混合碳源', 'MEIII': '混合碳源',
    'CI': '对照', 'CII': '对照',
}

POLLUTANT_TYPE = {
    'MI': 'Cr+AN', 'MII': 'Cr', 'MIII': 'AN',
    'EI': 'Cr+AN', 'EII': 'Cr', 'EIII': 'AN',
    'MEI': 'Cr+AN', 'MEII': 'Cr', 'MEIII': 'AN',
    'CI': 'Cr+AN', 'CII': 'Cr',
}

# ── 颜色方案（使用系统色板扩展）──
_COLORS = list(TABLEAU_10) + ['#1f77b4', '#ff7f0e']  # extend for 11 groups
GROUP_COLORS = dict(zip(ALL_GROUPS, _COLORS[:len(ALL_GROUPS)]))

# ══════════════════════════════════════
# 2. 解析各指标数据
# ══════════════════════════════════════

def find_group_cols(row_num):
    groups = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row_num, column=c).value
        if v and isinstance(v, str) and v.strip() in ALL_GROUPS:
            groups[v.strip()] = c
    return groups

def read_paired(start_row, end_row, groups, time_col=1):
    times = []
    for r in range(start_row, end_row + 1):
        t = cell(r, time_col)
        if not np.isnan(t):
            times.append((r, t))
    result = {}
    for name, col in groups.items():
        vals = []
        for r, t in times:
            v1, v2 = cell(r, col), cell(r, col + 1)
            vals.append({'time': t, 'rep1': v1, 'rep2': v2, 'mean': np.nanmean([v1, v2])})
        result[name] = pd.DataFrame(vals)
    return result

def read_fe(start_row, end_row, groups, time_col=1):
    times = []
    for r in range(start_row, end_row + 1):
        t = cell(r, time_col)
        if not np.isnan(t):
            times.append((r, t))
    result = {}
    for name, col in groups.items():
        vals = []
        for r, t in times:
            vals.append({
                'time': t,
                'Fe2_mean': np.nanmean([cell(r, col), cell(r, col+1)]),
                'Fe3_mean': np.nanmean([cell(r, col+2), cell(r, col+3)]),
                'TFe_mean': np.nanmean([cell(r, col+4), cell(r, col+5)]),
            })
        result[name] = pd.DataFrame(vals)
    return result

def extract_ph_do_orp(group_header_row):
    groups = find_group_cols(group_header_row)
    times = []
    for r in range(7, 16):
        t = cell(r, 1)
        if not np.isnan(t):
            times.append((r, t))
    ph, do_, orp = {}, {}, {}
    for name, col in groups.items():
        ph_v, do_v, orp_v = [], [], []
        for r, t in times:
            ph_v.append({'time': t, 'mean': np.nanmean([cell(r, col), cell(r, col+1)])})
            do_v.append({'time': t, 'mean': np.nanmean([cell(r, col+2), cell(r, col+3)])})
            orp_v.append({'time': t, 'mean': np.nanmean([cell(r, col+4), cell(r, col+5)])})
        ph[name] = pd.DataFrame(ph_v)
        do_[name] = pd.DataFrame(do_v)
        orp[name] = pd.DataFrame(orp_v)
    return ph, do_, orp

# ── Fe列位置（每组6列：Fe(II)×2, Fe(III)×2, TFe×2）──
FE_GROUPS_MAP = {
    'MI': 3, 'MII': 9, 'MIII': 15, 'EI': 21, 'EII': 27, 'EIII': 33,
    'MEI': 39, 'MEII': 45, 'MEIII': 51, 'CI': 57, 'CII': 63,
}

# ── 解析全部数据 ──
ph_data, do_data, orp_data = extract_ph_do_orp(5)
cr_data = read_paired(31, 42, find_group_cols(30))
tcr_data = read_paired(53, 57, find_group_cols(52))
tcr_data.pop('MEII', None)
toc_data = read_paired(65, 71, find_group_cols(64))
fe_data1 = read_fe(89, 95, FE_GROUPS_MAP)
fe_data2 = read_fe(111, 117, FE_GROUPS_MAP)
an_data = read_paired(133, 142, find_group_cols(132))

print("[OK] Data loaded successfully")

# ══════════════════════════════════════
# 3. 构建长格式 DataFrame（供 StatisticalAnalyzer 使用）
# ══════════════════════════════════════

def build_long_df(data_dict, indicator_name, groups=None):
    """将分组数据转为长格式 DataFrame"""
    if groups is None:
        groups = list(data_dict.keys())
    records = []
    for g in groups:
        if g not in data_dict:
            continue
        df = data_dict[g]
        for _, row in df.iterrows():
            if 'mean' in row and not np.isnan(row['mean']):
                records.append({
                    'group': g,
                    'time': row['time'],
                    indicator_name: row['mean'],
                    'carbon_source': CARBON_SOURCE.get(g, ''),
                    'pollutant_type': POLLUTANT_TYPE.get(g, ''),
                })
    return pd.DataFrame(records)

# 构建综合长格式表（用于相关性分析）
def build_comprehensive_long(groups):
    records = []
    for g in groups:
        if g not in cr_data:
            continue
        cr_df = cr_data[g]
        for _, row in cr_df.iterrows():
            t = row['time']
            rec = {
                'group': g, 'time': t,
                'Cr_VI': row['mean'] if not np.isnan(row['mean']) else np.nan,
                'carbon_source': CARBON_SOURCE[g],
                'pollutant_type': POLLUTANT_TYPE[g],
            }
            # AN
            if g in an_data:
                an_df = an_data[g]
                m = an_df[an_df['time'] == t]
                rec['AN'] = m['mean'].values[0] if len(m) > 0 and not np.isnan(m['mean'].values[0]) else np.nan
            else:
                rec['AN'] = np.nan
            # pH/DO/ORP
            for name, d in [('pH', ph_data), ('DO', do_data), ('ORP', orp_data)]:
                if g in d:
                    m = d[g][d[g]['time'] == t]
                    rec[name] = m['mean'].values[0] if len(m) > 0 else np.nan
            # TOC
            if g in toc_data:
                m = toc_data[g][toc_data[g]['time'] == t]
                rec['TOC'] = m['mean'].values[0] if len(m) > 0 else np.nan
            # Fe
            if g in fe_data2:
                m = fe_data2[g][fe_data2[g]['time'] == t]
                if len(m) > 0:
                    rec['Fe_II'] = m['Fe2_mean'].values[0]
                    rec['TFe'] = m['TFe_mean'].values[0]
            records.append(rec)
    return pd.DataFrame(records)

# ══════════════════════════════════════
# 4. 绘图函数（使用学术AI系统）
# ══════════════════════════════════════

def plot_time_series(data_dict, groups, title, ylabel, filename, c_groups=None):
    """时间序列图（学术样式）"""
    if c_groups is None:
        c_groups = C_GROUPS
    fig, ax = plt.subplots(figsize=(8, 5))
    # 对照组（虚线）
    for g in c_groups:
        if g in data_dict:
            df = data_dict[g]
            valid = df.dropna(subset=['mean'])
            ax.plot(valid['time'], valid['mean'], '--', color=GROUP_COLORS[g],
                   label=LABELS[g], linewidth=1.5, alpha=0.7)
    # 实验组（实线+标记）
    markers = ['o', 's', '^', 'D', 'v', 'P']
    for i, g in enumerate(groups):
        if g in data_dict:
            df = data_dict[g]
            valid = df.dropna(subset=['mean'])
            ax.plot(valid['time'], valid['mean'], '-o', color=GROUP_COLORS[g],
                   label=LABELS[g], linewidth=2, markersize=5,
                   markerfacecolor='white', markeredgecolor=GROUP_COLORS[g],
                   markeredgewidth=1.5)
    ax.set_xlabel('时间 (天)', fontproperties=CN_FONT_PROP, fontsize=12)
    ax.set_ylabel(ylabel, fontproperties=CN_FONT_PROP, fontsize=12)
    ax.set_title(title, fontproperties=CN_FONT_PROP_BOLD, fontsize=14)
    ax.legend(loc='best', fontsize=8, framealpha=0.9, prop=CN_FONT_PROP)
    ax.grid(True, alpha=0.3)
    ax.set_xlim(left=0)
    plt.tight_layout()
    save_figure(fig, filename, OUT_DIR, formats=['png'])
    plt.close()

def plot_correlation_heatmap(df, corr_cols, title, filename):
    """相关性热力图（下三角 + 显著性星号）"""
    valid = df[corr_cols].dropna()
    if len(valid) < 5:
        print(f"  [SKIP] {filename}: insufficient data ({len(valid)} rows)")
        return
    analyzer = StatisticalAnalyzer(valid)
    corr_matrix, p_matrix = analyzer.correlation_analysis(method='pearson', cols=corr_cols)
    n = len(corr_cols)
    fig, ax = plt.subplots(figsize=(max(8, n * 1.1), max(7, n * 1.0)))
    # 下三角掩码
    mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=0)
    import seaborn as sns
    sns.heatmap(corr_matrix, mask=mask, annot=False, fmt='', cmap='RdBu_r',
                center=0, vmin=-1, vmax=1, ax=ax, square=True,
                linewidths=0.5, cbar_kws={'shrink': 0.8, 'label': 'Pearson r'})
    # 在每个下三角格子里写 r值 + 显著性星号
    for i in range(n):
        for j in range(i + 1):  # 下三角（含对角线）
            r = corr_matrix.iloc[i, j]
            p = p_matrix.iloc[i, j]
            stars = significance_stars(p)
            # 对角线只写变量名
            if i == j:
                ax.text(j + 0.5, i + 0.5, f'{r:.2f}', ha='center', va='center',
                       fontsize=9, fontweight='bold', color='black')
            else:
                # r值
                txt = f'{r:.2f}'
                ax.text(j + 0.5, i + 0.4, txt, ha='center', va='center',
                       fontsize=9, color='black')
                # 显著性星号（红色加粗）
                if stars != 'n.s.':
                    ax.text(j + 0.5, i + 0.7, stars, ha='center', va='center',
                           fontsize=10, color='red', fontweight='bold')
                else:
                    ax.text(j + 0.5, i + 0.7, 'n.s.', ha='center', va='center',
                           fontsize=7, color='gray')
    # 轴标签
    ax.set_xticklabels(corr_cols, rotation=45, ha='right', fontsize=10)
    ax.set_yticklabels(corr_cols, rotation=0, fontsize=10)
    ax.set_title(title, fontproperties=CN_FONT_PROP_BOLD, fontsize=13, pad=15)
    # 添加样本量信息
    ax.text(0.01, -0.05, f'n = {len(valid)}', transform=ax.transAxes,
           fontsize=9, color='gray', va='top')
    plt.tight_layout()
    save_figure(fig, filename, OUT_DIR, formats=['png'])
    plt.close()

def plot_coexist_compare(data_dict, group_both, group_single, pollutant, title, filename):
    """共存影响对比图"""
    fig, ax = plt.subplots(figsize=(8, 5))
    for g in C_GROUPS:
        if g in data_dict:
            df = data_dict[g].dropna(subset=['mean'])
            ax.plot(df['time'], df['mean'], '--', color=GROUP_COLORS[g],
                   label=LABELS[g], linewidth=1.5, alpha=0.7)
    if group_both in data_dict:
        df = data_dict[group_both].dropna(subset=['mean'])
        ax.plot(df['time'], df['mean'], '-o', color=GROUP_COLORS[group_both],
               label=f'{LABELS[group_both]} (含两种污染物)', linewidth=2, markersize=5,
               markerfacecolor='white', markeredgecolor=GROUP_COLORS[group_both], markeredgewidth=1.5)
    if group_single in data_dict:
        df = data_dict[group_single].dropna(subset=['mean'])
        ax.plot(df['time'], df['mean'], '--s', color=GROUP_COLORS[group_single],
               label=f'{LABELS[group_single]} (仅含{pollutant})', linewidth=2, markersize=5)
    ax.set_xlabel('时间 (天)', fontproperties=CN_FONT_PROP, fontsize=12)
    ax.set_ylabel(f'{pollutant} 浓度 (mg/L)', fontproperties=CN_FONT_PROP, fontsize=12)
    ax.set_title(title, fontproperties=CN_FONT_PROP_BOLD, fontsize=14)
    ax.legend(loc='best', fontsize=8, framealpha=0.9, prop=CN_FONT_PROP)
    ax.grid(True, alpha=0.3)
    ax.set_xlim(left=0)
    plt.tight_layout()
    save_figure(fig, filename, OUT_DIR, formats=['png'])
    plt.close()

def print_stats(data_dict, groups, name):
    print(f"\n--- {name} ---")
    for g in groups + C_GROUPS:
        if g in data_dict:
            df = data_dict[g]
            if 'mean' in df.columns:
                valid = df.dropna(subset=['mean'])
                if len(valid) == 0:
                    continue
                init, final = valid['mean'].iloc[0], valid['mean'].iloc[-1]
                removal = (init - final) / init * 100 if init > 0 else 0
                print(f"  {g}: {init:.2f} -> {final:.2f} (removal {removal:.1f}%)")


# ══════════════════════════════════════
# 5. M组分析
# ══════════════════════════════════════
print("\n=== M组 (糖浆) ===")
plot_time_series(cr_data, M_GROUPS, 'M组（糖浆）Cr(VI)浓度变化趋势', 'Cr(VI) (mg/L)', 'M组_CrVI趋势')
plot_time_series(an_data, ['MI', 'MIII'], 'M组（糖浆）AN浓度变化趋势', 'AN (mg/L)', 'M组_AN趋势')
plot_time_series(ph_data, M_GROUPS, 'M组（糖浆）pH变化趋势', 'pH', 'M组_pH趋势')
plot_time_series(do_data, M_GROUPS, 'M组（糖浆）DO变化趋势', 'DO (mg/L)', 'M组_DO趋势')
plot_time_series(orp_data, M_GROUPS, 'M组（糖浆）ORP变化趋势', 'ORP (mV)', 'M组_ORP趋势')
plot_time_series(toc_data, M_GROUPS, 'M组（糖浆）TOC变化趋势', 'TOC (mg/L)', 'M组_TOC趋势')
plot_time_series({g: fe_data2[g].rename(columns={'Fe2_mean': 'mean'}) for g in fe_data2 if g in M_GROUPS},
    M_GROUPS, 'M组（糖浆）液相Fe(II)变化趋势', 'Fe(II) (ug/L)', 'M组_FeII趋势')
plot_time_series({g: fe_data2[g].rename(columns={'TFe_mean': 'mean'}) for g in fe_data2 if g in M_GROUPS},
    M_GROUPS, 'M组（糖浆）液相TFe变化趋势', 'TFe (ug/L)', 'M组_TFe趋势')
plot_time_series(tcr_data, ['MI', 'MII'], 'M组（糖浆）总铬浓度变化趋势', 'TCr (mg/L)', 'M组_TCr趋势')

# 相关性
df_m = build_comprehensive_long(M_GROUPS)
plot_correlation_heatmap(df_m, ['Cr_VI', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe'],
    'M组 Cr(VI)与环境因子Pearson相关性', 'M组_CrVI相关性')
plot_correlation_heatmap(df_m, ['AN', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe'],
    'M组 AN与环境因子Pearson相关性', 'M组_AN相关性')

# 共存影响
plot_coexist_compare(cr_data, 'MI', 'MII', 'Cr(VI)',
    'M组 Cr(VI)共存影响：MI(含AN) vs MII(不含AN)', 'M组_CrVI共存影响')
plot_coexist_compare(an_data, 'MI', 'MIII', 'AN',
    'M组 AN共存影响：MI(含Cr) vs MIII(不含Cr)', 'M组_AN共存影响')


# ══════════════════════════════════════
# 6. E组分析
# ══════════════════════════════════════
print("\n=== E组 (乳化植物油) ===")
plot_time_series(cr_data, E_GROUPS, 'E组（乳化植物油）Cr(VI)浓度变化趋势', 'Cr(VI) (mg/L)', 'E组_CrVI趋势')
plot_time_series(an_data, ['EI', 'EIII'], 'E组（乳化植物油）AN浓度变化趋势', 'AN (mg/L)', 'E组_AN趋势')
plot_time_series(ph_data, E_GROUPS, 'E组（乳化植物油）pH变化趋势', 'pH', 'E组_pH趋势')
plot_time_series(do_data, E_GROUPS, 'E组（乳化植物油）DO变化趋势', 'DO (mg/L)', 'E组_DO趋势')
plot_time_series(orp_data, E_GROUPS, 'E组（乳化植物油）ORP变化趋势', 'ORP (mV)', 'E组_ORP趋势')
plot_time_series(toc_data, E_GROUPS, 'E组（乳化植物油）TOC变化趋势', 'TOC (mg/L)', 'E组_TOC趋势')
plot_time_series({g: fe_data2[g].rename(columns={'Fe2_mean': 'mean'}) for g in fe_data2 if g in E_GROUPS},
    E_GROUPS, 'E组（乳化植物油）液相Fe(II)变化趋势', 'Fe(II) (ug/L)', 'E组_FeII趋势')
plot_time_series({g: fe_data2[g].rename(columns={'TFe_mean': 'mean'}) for g in fe_data2 if g in E_GROUPS},
    E_GROUPS, 'E组（乳化植物油）液相TFe变化趋势', 'TFe (ug/L)', 'E组_TFe趋势')
plot_time_series(tcr_data, ['EI', 'EII'], 'E组（乳化植物油）总铬浓度变化趋势', 'TCr (mg/L)', 'E组_TCr趋势')

df_e = build_comprehensive_long(E_GROUPS)
plot_correlation_heatmap(df_e, ['Cr_VI', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe'],
    'E组 Cr(VI)与环境因子Pearson相关性', 'E组_CrVI相关性')
plot_correlation_heatmap(df_e, ['AN', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe'],
    'E组 AN与环境因子Pearson相关性', 'E组_AN相关性')

plot_coexist_compare(cr_data, 'EI', 'EII', 'Cr(VI)',
    'E组 Cr(VI)共存影响：EI(含AN) vs EII(不含AN)', 'E组_CrVI共存影响')
plot_coexist_compare(an_data, 'EI', 'EIII', 'AN',
    'E组 AN共存影响：EI(含Cr) vs EIII(不含Cr)', 'E组_AN共存影响')


# ══════════════════════════════════════
# 7. ME组分析
# ══════════════════════════════════════
print("\n=== ME组 (糖浆+乳化植物油) ===")
plot_time_series(cr_data, ME_GROUPS, 'ME组（混合碳源）Cr(VI)浓度变化趋势', 'Cr(VI) (mg/L)', 'ME组_CrVI趋势')
plot_time_series(an_data, ['MEI', 'MEIII'], 'ME组（混合碳源）AN浓度变化趋势', 'AN (mg/L)', 'ME组_AN趋势')
plot_time_series(ph_data, ME_GROUPS, 'ME组（混合碳源）pH变化趋势', 'pH', 'ME组_pH趋势')
plot_time_series(do_data, ME_GROUPS, 'ME组（混合碳源）DO变化趋势', 'DO (mg/L)', 'ME组_DO趋势')
plot_time_series(orp_data, ME_GROUPS, 'ME组（混合碳源）ORP变化趋势', 'ORP (mV)', 'ME组_ORP趋势')
plot_time_series(toc_data, ME_GROUPS, 'ME组（混合碳源）TOC变化趋势', 'TOC (mg/L)', 'ME组_TOC趋势')
plot_time_series({g: fe_data2[g].rename(columns={'Fe2_mean': 'mean'}) for g in fe_data2 if g in ME_GROUPS},
    ME_GROUPS, 'ME组（混合碳源）液相Fe(II)变化趋势', 'Fe(II) (ug/L)', 'ME组_FeII趋势')
plot_time_series({g: fe_data2[g].rename(columns={'TFe_mean': 'mean'}) for g in fe_data2 if g in ME_GROUPS},
    ME_GROUPS, 'ME组（混合碳源）液相TFe变化趋势', 'TFe (ug/L)', 'ME组_TFe趋势')
plot_time_series(tcr_data, ['MEI'], 'ME组（混合碳源）总铬浓度变化趋势', 'TCr (mg/L)', 'ME组_TCr趋势')

df_me = build_comprehensive_long(ME_GROUPS)
plot_correlation_heatmap(df_me, ['Cr_VI', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe'],
    'ME组 Cr(VI)与环境因子Pearson相关性', 'ME组_CrVI相关性')
plot_correlation_heatmap(df_me, ['AN', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe'],
    'ME组 AN与环境因子Pearson相关性', 'ME组_AN相关性')

plot_coexist_compare(cr_data, 'MEI', 'MEII', 'Cr(VI)',
    'ME组 Cr(VI)共存影响：MEI(含AN) vs MEII(不含AN)', 'ME组_CrVI共存影响')
plot_coexist_compare(an_data, 'MEI', 'MEIII', 'AN',
    'ME组 AN共存影响：MEI(含Cr) vs MEIII(不含Cr)', 'ME组_AN共存影响')


# ══════════════════════════════════════
# 8. 综合对比图
# ══════════════════════════════════════
print("\n=== 综合对比 ===")
fig, axes = plt.subplots(1, 3, figsize=(18, 5))

for idx, (subtitle, groups) in enumerate([
    ('仅含Cr(VI)', ['MII', 'EII', 'MEII']),
    ('Cr(VI)+AN共存', ['MI', 'EI', 'MEI']),
]):
    ax = axes[idx]
    for g in C_GROUPS:
        if g in cr_data:
            df = cr_data[g].dropna(subset=['mean'])
            ax.plot(df['time'], df['mean'], '--', color=GROUP_COLORS[g], label='对照', alpha=0.7)
    for g in groups:
        if g in cr_data:
            df = cr_data[g].dropna(subset=['mean'])
            ax.plot(df['time'], df['mean'], '-o', color=GROUP_COLORS[g],
                   label=LABELS[g], linewidth=2, markersize=4,
                   markerfacecolor='white', markeredgecolor=GROUP_COLORS[g], markeredgewidth=1.5)
    ax.set_xlabel('时间 (天)', fontproperties=CN_FONT_PROP)
    ax.set_ylabel('Cr(VI) (mg/L)', fontproperties=CN_FONT_PROP)
    ax.set_title(f'Cr(VI)下降趋势 ({subtitle})', fontproperties=CN_FONT_PROP_BOLD, fontsize=12)
    ax.legend(fontsize=7, prop=CN_FONT_PROP)
    ax.grid(True, alpha=0.3)

ax = axes[2]
for g in C_GROUPS:
    if g in an_data:
        df = an_data[g].dropna(subset=['mean'])
        ax.plot(df['time'], df['mean'], '--', color=GROUP_COLORS[g], label='对照', alpha=0.7)
for g in ['MI', 'MIII', 'EI', 'EIII', 'MEI', 'MEIII']:
    if g in an_data:
        df = an_data[g].dropna(subset=['mean'])
        ax.plot(df['time'], df['mean'], '-o', color=GROUP_COLORS[g],
               label=LABELS[g], linewidth=1.5, markersize=4,
               markerfacecolor='white', markeredgecolor=GROUP_COLORS[g], markeredgewidth=1)
ax.set_xlabel('时间 (天)', fontproperties=CN_FONT_PROP)
ax.set_ylabel('AN (mg/L)', fontproperties=CN_FONT_PROP)
ax.set_title('AN下降趋势 (各组对比)', fontproperties=CN_FONT_PROP_BOLD, fontsize=12)
ax.legend(fontsize=6, ncol=2, prop=CN_FONT_PROP)
ax.grid(True, alpha=0.3)

plt.suptitle('不同碳源对Cr(VI)和AN去除效果对比', fontproperties=CN_FONT_PROP_BOLD, fontsize=15, y=1.02)
plt.tight_layout()
save_figure(fig, '综合对比_碳源效果', OUT_DIR, formats=['png'])
plt.close()


# ══════════════════════════════════════
# 9. Fe综合图
# ══════════════════════════════════════
print("\n=== Fe综合图 ===")
fig, axes = plt.subplots(2, 3, figsize=(18, 10))
for idx, (name, groups) in enumerate([('M', M_GROUPS), ('E', E_GROUPS), ('ME', ME_GROUPS)]):
    for row, (fe_key, ylabel, title_suffix) in enumerate([
        ('Fe2_mean', 'Fe(II) (ug/L)', '液相Fe(II)'),
        ('TFe_mean', 'TFe (ug/L)', '液相TFe'),
    ]):
        ax = axes[row, idx]
        for g in C_GROUPS:
            if g in fe_data2:
                df = fe_data2[g].dropna(subset=[fe_key])
                ax.plot(df['time'], df[fe_key], '--', color=GROUP_COLORS[g], label='对照', alpha=0.7)
        for g in groups:
            if g in fe_data2:
                df = fe_data2[g].dropna(subset=[fe_key])
                ax.plot(df['time'], df[fe_key], '-o', color=GROUP_COLORS[g],
                       label=LABELS[g], markersize=4, markerfacecolor='white',
                       markeredgecolor=GROUP_COLORS[g], markeredgewidth=1)
        ax.set_xlabel('时间 (天)', fontproperties=CN_FONT_PROP)
        ax.set_ylabel(ylabel, fontproperties=CN_FONT_PROP)
        ax.set_title(f'{name}组 {title_suffix}', fontproperties=CN_FONT_PROP_BOLD)
        ax.legend(fontsize=7, prop=CN_FONT_PROP)
        ax.grid(True, alpha=0.3)

plt.suptitle('液相铁含量变化趋势', fontproperties=CN_FONT_PROP_BOLD, fontsize=15)
plt.tight_layout()
save_figure(fig, 'Fe含量综合趋势', OUT_DIR, formats=['png'])
plt.close()


# ══════════════════════════════════════
# 10. 统计摘要
# ══════════════════════════════════════
print("\n=== 统计摘要 ===")
print_stats(cr_data, M_GROUPS, "M组 Cr(VI)")
print_stats(cr_data, E_GROUPS, "E组 Cr(VI)")
print_stats(cr_data, ME_GROUPS, "ME组 Cr(VI)")
print_stats(an_data, ['MI', 'MIII'], "M组 AN")
print_stats(an_data, ['EI', 'EIII'], "E组 AN")
print_stats(an_data, ['MEI', 'MEIII'], "ME组 AN")


# ══════════════════════════════════════
# 11. 全量相关性大表
# ══════════════════════════════════════
print("\n=== 全量相关性分析 ===")
df_all = build_comprehensive_long(ALL_GROUPS)
corr_cols = ['Cr_VI', 'AN', 'pH', 'DO', 'ORP', 'TOC', 'Fe_II', 'TFe']
plot_correlation_heatmap(df_all, corr_cols,
    '全部组别 污染物与环境因子Pearson相关性', '全量相关性热力图')


print(f"\n[DONE] All figures saved to: {OUT_DIR}")
